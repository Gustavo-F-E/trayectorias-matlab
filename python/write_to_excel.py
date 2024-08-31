import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def write_to_excel(excel_path, sheet_name, df):
    try:
        # Cargar el libro si existe, o crear uno nuevo si no
        if os.path.exists(excel_path):
            book = load_workbook(excel_path)
        else:
            book = Workbook()
        
        # Si la hoja ya existe, eliminarla
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        
        # Crear una nueva hoja
        sheet = book.create_sheet(title=sheet_name)
        
        # Escribir el DataFrame en la hoja
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
        
        # Guardar el archivo
        book.save(excel_path)
        print(f'Se escribió la hoja "{sheet_name}" en el archivo Excel.')
    except PermissionError as pe:
        print(f'Error de permisos al escribir en el archivo Excel: {pe}')
    except Exception as e:
        print(f'Error al escribir en el archivo Excel: {e}')