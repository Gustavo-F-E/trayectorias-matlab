import openpyxl

def update_excel_from_dict(excel_path, sheet_name, var_ini):
    try:
        # Cargar el archivo Excel y la hoja especificada
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb[sheet_name]

        # Recorrer todas las filas en la columna 'A' para buscar las claves
        for row in sheet.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            key = cell.value
            if key in var_ini:
                # Actualizar el valor en la columna 'B'
                row_num = cell.row
                sheet.cell(row=row_num, column=2, value=var_ini[key])

        # Guardar los cambios en el archivo Excel
        wb.save(excel_path)
        print(f'Archivo "{excel_path}" actualizado correctamente.')

    except Exception as e:
        print(f'Error al actualizar el archivo Excel: {e}')

# Diccionario de ejemplo
var_ini = {
    'b': 6,
    'rim': 75,
    'rfm': 75,
    'zim': -500,
    'zi': 0,
    'zfg': 300,
    'zfm': 1000,
    'alfainicio': 45,
    'lambda1': 0.14
}